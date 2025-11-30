import os
import sys
import django
from pathlib import Path
import openpyxl
from datetime import datetime
from django.utils.timezone import make_aware

# === Configuración Django ===
BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.append(str(BASE_DIR))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'emcali_app.settings')
django.setup()

# === Importar modelos ===
from mantenimiento.models import Nodo, Orden, InformeReco, HistoricoReco

# Ruta al Excel
RUTA_EXCEL = BASE_DIR / "mantenimiento" / "scripts" / "trabajosrecos.xlsx"


def str_to_datetime(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return make_aware(valor)

    formatos = [
        "%m/%d/%y %H:%M",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M",
    ]

    for f in formatos:
        try:
            return make_aware(datetime.strptime(str(valor), f))
        except:
            pass

    return None


def safe_datetime(valor):
    dt = str_to_datetime(valor)
    if dt is None:
        return make_aware(datetime.now())
    return dt


def safe_float(valor, minimo=-180, maximo=180):
    if valor is None:
        return 0

    try:
        valor = str(valor).replace(",", ".")
        num = float(valor)
        if minimo <= num <= maximo:
            return num
        return 0
    except:
        return 0


def cargar_informes():
    wb = openpyxl.load_workbook(RUTA_EXCEL)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))

        nodo_id = str(data.get("NODO")).strip()

        # Buscar nodo
        try:
            nodo = Nodo.objects.get(nodo=nodo_id)
        except Nodo.DoesNotExist:
            print(f"❌ Nodo no encontrado: {nodo_id}")
            continue

        descripcion_excel = data.get("DESCRIPCION DE LA ORDEN")
        descripcion_excel = "" if descripcion_excel is None else str(descripcion_excel).strip()

        # === CREAR ORDEN ÚNICA POR FILA ===
        n_orden_unico = f"{nodo_id}-{index}"

        orden, creada_orden = Orden.objects.get_or_create(
            n_orden=n_orden_unico,
            id_nodo=nodo,
            defaults={
                "descripcion": descripcion_excel,
                "fecha": datetime.now().date(),
                "estado_orden": "revisada",
                "tipo": "reco",
                "creada_por_id": 1,
                "asignada_a_id": 1,
            }
        )

        if creada_orden:
            print(f"🆕 Orden creada: {n_orden_unico}")
        else:
            print(f"➡ Orden existente (no debería pasar): {n_orden_unico}")

        # Crear INFORME
        informe = InformeReco.objects.create(
            orden=orden,
            inicio_actividad=safe_datetime(data.get("INICIO ACTIVIDAD")),
            hora_cierre=safe_datetime(data.get("FIN ACTIVIDAD")),
            aviso_mante=str(data.get("AVISO MANTENIMIENTO")).upper() == "TRUE",
            aviso_telco=str(data.get("AVISO TELCO")).upper() == "TRUE",
            detalle_mante=data.get("DESCRIPCION AVISO") or "",
            detalle_telco=data.get("DESCRIPCION AVISO") or "",
            falla_identificada=data.get("FALLA IDENTIFICADA") or "",
            causa_falla=data.get("CAUSA DE FALLA") or "",
            trabajo_realizado=data.get("TRABAJO REALIZADO") or "",
            latitud=safe_float(data.get("LATITUD"), minimo=-90, maximo=90),
            longitud=safe_float(data.get("LONGITUD"), minimo=-180, maximo=180),
            id_fotos="",  # Fotos ignoradas
        )

        print(f"✔ Informe creado # {informe.id}")

        # Crear histórico por cada informe
        histo = HistoricoReco.objects.create()
        histo.informes.add(informe)
        print(f"📘 Histórico creado para informe #{informe.id}\n")


if __name__ == "__main__":
    cargar_informes()
